#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
工程核心文件 - 报销自动化工作流程控制
负责管理各个阶段之间的跳转操作和业务逻辑
"""

from typing import Dict, List, Optional, Any, Set
from enum import Enum
import os
import json
import requests
import textwrap
import traceback

# 可选依赖（首次用到映射表时才需要）
try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None  # 延迟报错
    load_workbook = None

class WorkflowStage(Enum):
    """工作流程阶段枚举"""
    LOGIN = "login"
    PROJECT = "project"
    EXPENSE = "expense"
    PERSONNEL = "personnel"
    APPOINTMENT = "appointment"
    TRAVEL = "travel"
    LABOR = "labor"

class WorkflowCore:
    """工作流程核心控制器"""
    
    def __init__(self):
        """初始化工作流程控制器"""
        self._init_stage_transitions()
        self.ollama_base_url: str = os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434")
        self.ollama_model: str = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")
        self.mapping_excel_path: str = os.path.join(os.path.dirname(__file__), "field_type_mapping.xlsx")
    
    def _init_stage_transitions(self):
        """初始化各阶段跳转操作说明"""
        
        # 1. 登录阶段跳转操作说明
        self.LOGIN_TRANSITION = textwrap.dedent("""
            将ID为checkcodeImg的验证码图片保存至C:\\Users\\FH\\source\\repos\\Auto Finan\\Auto Finan\\LLM_Integration目录下，命名为example.jpg
            运行C:\\Users\\FH\\source\\repos\\Auto Finan\\Auto Finan\\LLM_Integration目录下的OCR.py，得到读取的验证码
            输入验证码
            点击id为zhLogin的登录按钮
            点击网上预约报账按钮
            点击申请报销单按钮
            点击已阅读并同意按钮
        """).strip()
        
        # 2. 项目信息阶段跳转操作说明
        self.PROJECT_TRANSITION = textwrap.dedent("""
            点击下一步按钮
        """).strip()
        
        # 3. 报销科目信息阶段跳转操作说明
        self.EXPENSE_TRANSITION = textwrap.dedent("""
            点击下一步按钮
        """).strip()
        
        # 4. 报销人员信息阶段跳转操作说明
        self.PERSONNEL_TRANSITION = textwrap.dedent("""
            点击下一步按钮
        """).strip()
        
        # 5. 预约时间阶段跳转操作说明
        self.APPOINTMENT_TRANSITION = textwrap.dedent("""
            点击预约按钮
            点击打印确认单按钮
            调用test_mouse_keyboard.py，执行一个python自动点击的脚本，脚本的第一个参数为保存路径，第二个参数为保存文件名，请你以当前页面中的信息，以报销单号-项目号-金额的格式，输入第二个参数
            等待刚刚运行的脚本运行完毕
            点击返回按钮
            重命名当前读取的提示词文件，将未预约改成已预约
        """).strip()
        
        # 6. 差旅信息跳转操作说明
        self.TRAVEL_TRANSITION = textwrap.dedent("""
            在填写完信息后，请你再检查一下人员信息中的人员类型有没有正确填写，网页有的时候会覆盖之前的填写
            点击下一步按钮
            等待页面跳转
        """).strip()
        
        # 7. 劳务信息跳转操作说明
        self.LABOR_TRANSITION = textwrap.dedent("""
            点击下一步按钮
        """).strip()
        
        # 8. 单个人员信息填写后操作
        self.PERSONNEL_ITEM_POST_ACTION = textwrap.dedent("""
            点击提交按钮
            等待页面响应
        """).strip()
        
        # 9. 单个劳务酬金信息填写后操作
        self.LABOR_ITEM_POST_ACTION = textwrap.dedent("""
            点击确定按钮
            等待页面响应
        """).strip()

        # 10. 劳务信息(laborInfo)填写结束后，跳转到劳务人员(laborPerson)的操作说明
        self.LABORINFO_TO_LABORPERSON_TRANSITION = textwrap.dedent("""
            点击下一步按钮
        """).strip()

        # 11. 开始填写 laborPerson 之前的操作说明
        self.LABORPERSON_PRE_ACTION = textwrap.dedent("""
            点击单笔录入按钮
        """).strip()
    
    def get_transition_description(self, stage: WorkflowStage) -> str:
        """
        获取指定阶段的跳转操作说明
        
        Args:
            stage: 工作流程阶段
            
        Returns:
            str: 跳转操作说明字符串
        """
        transition_map = {
            WorkflowStage.LOGIN: self.LOGIN_TRANSITION,
            WorkflowStage.PROJECT: self.PROJECT_TRANSITION,
            WorkflowStage.EXPENSE: self.EXPENSE_TRANSITION,
            WorkflowStage.PERSONNEL: self.PERSONNEL_TRANSITION,
            WorkflowStage.APPOINTMENT: self.APPOINTMENT_TRANSITION,
            WorkflowStage.TRAVEL: self.TRAVEL_TRANSITION,
            WorkflowStage.LABOR: self.LABOR_TRANSITION
        }
        
        return transition_map.get(stage, "未知阶段")
    
    def get_all_transitions(self) -> Dict[str, str]:
        """
        获取所有阶段的跳转操作说明
        
        Returns:
            Dict[str, str]: 所有跳转操作说明的字典
        """
        return {
            "login": self.LOGIN_TRANSITION,
            "project": self.PROJECT_TRANSITION,
            "expense": self.EXPENSE_TRANSITION,
            "personnel": self.PERSONNEL_TRANSITION,
            "appointment": self.APPOINTMENT_TRANSITION,
            "travel": self.TRAVEL_TRANSITION,
            "labor": self.LABOR_TRANSITION
        }
    
    def print_all_transitions(self):
        """打印所有阶段的跳转操作说明"""
        print("=== 报销自动化工作流程 - 各阶段跳转操作说明 ===\n")
        
        stages = [
            ("1. 登录阶段", WorkflowStage.LOGIN),
            ("2. 项目信息阶段", WorkflowStage.PROJECT),
            ("3. 报销科目信息阶段", WorkflowStage.EXPENSE),
            ("4. 报销人员信息阶段", WorkflowStage.PERSONNEL),
            ("5. 预约时间阶段", WorkflowStage.APPOINTMENT),
            ("6. 差旅信息阶段", WorkflowStage.TRAVEL),
            ("7. 劳务信息阶段", WorkflowStage.LABOR)
        ]
        
        for stage_name, stage_enum in stages:
            print(f"{stage_name}")
            print(self.get_transition_description(stage_enum))
            print("-" * 80)

    # =========================
    # 新增：信息提取与字段类型映射
    # =========================

    def extract_form_json(self, user_input: str) -> Dict[str, Any]:
        """调用本地Qwen通过Ollama提取结构化JSON。

        Args:
            user_input: 自然语言输入

        Returns:
            提取到的JSON字典（失败时返回带error的字典）
        """
        prompt = self._build_extraction_prompt(user_input)
        try:
            resp = requests.post(
                f"{self.ollama_base_url}/api/generate",
                json={
                    "model": self.ollama_model,
                    "prompt": prompt,
                    "stream": False,
                    "options": {"temperature": 0.1, "top_p": 0.9, "max_tokens": 1200},
                },
                timeout=60,
            )
            if resp.status_code != 200:
                return {"error": f"LLM请求失败: {resp.status_code}", "detail": resp.text}
            data = resp.json()
            text = data.get("response", "").strip()
            # 直接JSON或代码块JSON
            try:
                if text.startswith('{'):
                    return json.loads(text)
            except Exception:
                pass
            # 从```json```代码块提取
            import re
            m = re.search(r"```json\s*(.*?)\s*```", text, re.DOTALL)
            if m:
                return json.loads(m.group(1).strip())
            # 兜底：返回原文
            return {"raw_response": text}
        except Exception as e:
            return {"error": f"LLM请求异常: {e}"}

    def _build_extraction_prompt(self, user_input: str) -> str:
        """与信息提取保持一致的提示词（与测试脚本语义一致，使用英文键名）。"""
        return (
            "你是一个专业的财务报销信息提取助手。请从用户输入的自然语言中提取报销相关信息，"
            "并严格输出JSON且只输出JSON，不要包含任何其他文字。\n\n"
            "数值规范：\n"
            "- 所有金额字段仅输出数字，不要带单位（例如 '500' 而非 '500元'）。涉及字段：expenses[].amount、personnel[].amount、labor[].amount、travel[].airfare、travel[].trainfare、travel[].otherTransport、travel[].accommodation。\n"
            "- 附件张数仅输出数字，不要带'张'（例如 '3' 而非 '3张'）。涉及字段：project.attachmentCount。\n"
            "- 布尔值字段输出true/false，不要输出中文。涉及字段：travel[].mealArranged、travel[].transportArranged。\n\n"
            "键名与结构：\n"
            "{\n"
            "  \"businessType\": \"业务大类\",\n"
            "  \"login\": {\n    \"username\": \"用户名\",\n    \"password\": \"密码\"\n  },\n"
            "  \"project\": {\n    \"projectNumber\": \"项目号\",\n    \"attachmentCount\": \"附件张数(数字)\",\n    \"paymentMethod\": \"支付方式\"\n  },\n"
            "  \"expenses\": [{\n    \"category\": \"科目类型\",\n    \"amount\": \"金额(数字)\"\n  }],\n"
            "  \"personnel\": [{\n    \"name\": \"姓名\",\n    \"ID\": \"学工号\",\n    \"bankCard\": \"银行卡信息\",\n    \"amount\": \"个人金额(数字)\"\n  }],\n"
            "  \"appointment\": {\n    \"date\": \"报销时间\",\n    \"location\": \"地点\"\n  },\n"
            "  \"travelPerson\": [{\n    \"ID\": \"出差人\",\n    \"name\": \"姓名\",\n    \"personType\": \"人员类型\",\n    \"workUnit\": \"工作单位\",\n    \"title\": \"职称\"\n  }],\n"
            "  \"travelExpenses\": [{\n    \"province\": \"省份\",\n    \"startTime\": \"起始时间\",\n    \"endTime\": \"结束时间\",\n    \"airfare\": \"飞机票(数字)\",\n    \"trainfare\": \"火车票(数字)\",\n    \"otherTransport\": \"其他交通费(数字)\",\n    \"accommodation\": \"住宿费(数字)\",\n    \"mealArranged\": \"是否安排伙食(true/false)\",\n    \"transportArranged\": \"是否安排交通(true/false)\"\n  }],\n"
            "  \"laborInfo\": {\n    \"personnelCategory\": \"人员类别\",\n    \"remunerationNature\": \"酬金性质\",\n    \"laborType\": \"劳务费类型\",\n    \"reason\": \"发放事由\",\n    \"remarks\": \"酬金信息备注\",\n    \"paymentStandard\": \"发放标准\",\n    \"startTime\": \"开始时间\",\n    \"endTime\": \"结束时间\"\n  },\n"
            "  \"laborPerson\": [{\n    \"employeeId\": \"工号/证件号\",\n    \"singleEntryAmount\": \"单笔录入金额(数字)\"\n  }]\n"
            "}\n\n"
            f"用户输入：{user_input}\n"
        )

    def collect_field_paths(self, data: Dict[str, Any]) -> List[str]:
        """从提取到的JSON中收集字段路径（用于生成/校验Excel映射）。

        使用点路径，数组用[]表示，如：expenses[].category
        """
        paths: Set[str] = set()

        def visit(node: Any, prefix: str = ""):
            if isinstance(node, dict):
                for k, v in node.items():
                    new_prefix = f"{prefix}.{k}" if prefix else k
                    visit(v, new_prefix)
            elif isinstance(node, list):
                # 标记为数组
                new_prefix = f"{prefix}[]" if prefix and not prefix.endswith("[]") else prefix
                if node:
                    # 取首元素推断结构
                    visit(node[0], new_prefix)
                else:
                    paths.add(new_prefix)
            else:
                paths.add(prefix)

        visit(data)
        return sorted(paths)

    def ensure_mapping_excel(self, field_paths: List[str]) -> str:
        """确保Excel映射存在；若不存在则创建。若存在则只补充缺失字段。

        Excel格式：第一列=元素名（字段路径），第二列=类型（i/c/r）
        """
        if Workbook is None or load_workbook is None:
            raise RuntimeError("请先安装依赖：pip install openpyxl")

        path = self.mapping_excel_path
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = "mapping"
            ws.append(["element", "type"])  # 表头
            for p in field_paths:
                ws.append([p, "i"])  # 默认输入框，后续可手工改为c/r
            wb.save(path)
            return path

        # 已存在：加载并补齐缺失字段，不覆盖已有配置
        wb = load_workbook(path)
        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                existing.add(str(row[0]))
        missing = [p for p in field_paths if p not in existing]
        for p in missing:
            ws.append([p, "i"])  # 默认输入框
        if missing:
            wb.save(path)
        return path

    def load_field_type_mapping(self) -> Dict[str, Dict[str, str]]:
        """读取Excel映射为字典：{ 字段路径: { type: i/c/r/d/'' , label: 中文名称 } }

        支持表头：第一列 json字段，第二列 标记，第三列 中文名称
        兼容旧版两列表头：element / type
        """
        if Workbook is None or load_workbook is None:
            raise RuntimeError("请先安装依赖：pip install openpyxl")
        path = self.mapping_excel_path
        if not os.path.exists(path):
            return {}
        wb = load_workbook(path)
        ws = wb.active
        # 解析表头
        headers_raw = [str(c or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = [h.lower() for h in headers_raw]
        # 头部字段名兼容
        def col_index(names):
            for n in names:
                key = n.lower()
                if key in headers:
                    return headers.index(key)
            return None

        key_idx = col_index(["json字段", "element", "字段", "key", "元素名"]) or 0
        type_idx = col_index(["标记", "type", "类型"]) or 1
        label_idx = col_index(["中文名称", "label", "名称"])  # 可为None

        mapping: Dict[str, Dict[str, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            key = str(row[key_idx]).strip() if key_idx is not None and row[key_idx] is not None else ""
            mark = str(row[type_idx]).strip().lower() if type_idx is not None and len(row) > type_idx and row[type_idx] is not None else ""
            label = str(row[label_idx]).strip() if label_idx is not None and len(row) > label_idx and row[label_idx] is not None else ""
            if not key:
                continue
            mapping[key] = {"type": mark, "label": label}
        return mapping

    def build_or_update_mapping_from_input(self, user_input: str) -> Dict[str, str]:
        """一体化流程：调用LLM提取→收集字段→创建/更新Excel→读取映射。

        Returns:
            映射字典 { 字段路径: { type, label } }
        """
        data = self.extract_form_json(user_input)
        if "error" in data:
            return data  # 直接返回错误信息
        # 兼容LLM非JSON返回
        if not isinstance(data, dict):
            return {"error": "LLM未返回JSON"}
        fields = self.collect_field_paths(data)
        self.ensure_mapping_excel(fields)
        return self.load_field_type_mapping()

    # =========================
    # 新增：根据JSON与Excel映射拼接Playwright提示词
    # =========================

    def _flatten_json(self, data: Any, prefix: str = "") -> List[tuple]:
        """展开JSON为 (路径, 值) 列表。数组使用具体索引以保留多项，后续匹配时会归一化为[]。"""
        items: List[tuple] = []
        if isinstance(data, dict):
            for k, v in data.items():
                new_prefix = f"{prefix}.{k}" if prefix else k
                items.extend(self._flatten_json(v, new_prefix))
        elif isinstance(data, list):
            for idx, v in enumerate(data):
                new_prefix = f"{prefix}[{idx}]" if prefix else f"[{idx}]"
                items.extend(self._flatten_json(v, new_prefix))
        else:
            items.append((prefix, data))
        return items

    @staticmethod
    def _normalize_array_path(path: str) -> str:
        """将路径中的具体索引标准化为[]，如 personnel[0].name -> personnel[].name"""
        import re
        return re.sub(r"\[\d+\]", "[]", path)

    def _format_segments(self, segments: List[str]) -> str:
        """将操作列表格式化为带序号的多行文本。"""
        numbered_segments = []
        for i, segment in enumerate(segments, 1):
            seg = str(segment).strip()
            if seg:
                numbered_segments.append(f"{i}. {seg}")
        return "\n".join(numbered_segments)

    def _get_stage_flow_config(self, business_type: str) -> List[Dict[str, str]]:
        """根据业务类型返回阶段配置。"""
        default_flow = [
            {"key": "login", "title": "登录阶段", "transition": self.LOGIN_TRANSITION},
            {"key": "project", "title": "项目信息阶段", "transition": self.PROJECT_TRANSITION},
            {"key": "expenses", "title": "科目信息阶段", "transition": self.EXPENSE_TRANSITION},
            {"key": "personnel", "title": "转卡信息阶段", "transition": self.PERSONNEL_TRANSITION},
            {"key": "appointment", "title": "预约阶段", "transition": self.APPOINTMENT_TRANSITION},
        ]
        travel_flow = [
            {"key": "login", "title": "登录阶段", "transition": self.LOGIN_TRANSITION},
            {"key": "project", "title": "项目信息阶段", "transition": self.PROJECT_TRANSITION},
            {"key": "travelPerson", "title": "出差人员阶段", "transition": ""},
            {"key": "travelExpenses", "title": "差旅费用阶段", "transition": self.TRAVEL_TRANSITION},
            {"key": "personnel", "title": "转卡信息阶段", "transition": self.PERSONNEL_TRANSITION},
            {"key": "appointment", "title": "预约阶段", "transition": self.APPOINTMENT_TRANSITION},
        ]
        labor_flow = [
            {"key": "login", "title": "登录阶段", "transition": self.LOGIN_TRANSITION},
            {"key": "project", "title": "项目信息阶段", "transition": self.PROJECT_TRANSITION},
            {"key": "laborInfo", "title": "劳务信息阶段", "transition": self.LABOR_TRANSITION},
            {"key": "laborPerson", "title": "劳务人员阶段", "transition": self.PERSONNEL_TRANSITION},
            {"key": "appointment", "title": "预约阶段", "transition": self.APPOINTMENT_TRANSITION},
        ]
        flow_map = {
            "报销业务": default_flow,
            "业务出差旅费": travel_flow,
            "酬金业务": labor_flow,
        }
        return flow_map.get(business_type, default_flow)

    def build_playwright_prompt_from_data(self, data: Dict[str, Any]) -> str:
        """根据提取的JSON数据与Excel映射，生成Playwright MCP提示词字符串。"""
        # 只读取现有映射，不创建或修改Excel文件
        mapping = self.load_field_type_mapping()
        parts: List[str] = []

        # 业务大类作为开头说明（如果有）
        business_type = data.get("businessType")
        if business_type:
            parts.append(f"业务大类：{business_type}。以下是需要执行的页面操作：")

        # 若为"报销业务"，按阶段顺序生成并在每阶段后追加跳转说明
        if (business_type or "").strip() == "报销业务":
            staged = self._build_prompt_reimbursement_flow(data, mapping)
            return staged
        
        # 若为"业务出差旅费"，按差旅费流程生成
        if (business_type or "").strip() == "业务出差旅费":
            staged = self._build_prompt_travel_flow(data, mapping)
            return staged
        
        # 若为"酬金业务"，按酬金流程生成
        if (business_type or "").strip() == "酬金业务":
            staged = self._build_prompt_labor_flow(data, mapping)
            return staged

        for path, value in self._flatten_json(data):
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            norm_path = self._normalize_array_path(path)
            # 特殊处理：expenses[].amount → 使用同项的category作为控件名称
            if norm_path == "expenses[].amount":
                # 从原始路径提取索引，找到对应category
                import re
                m = re.search(r"expenses\[(\d+)\]\.amount", path)
                if m:
                    idx = int(m.group(1))
                    try:
                        item = (data.get("expenses") or [])[idx]
                        cat = item.get("category") if isinstance(item, dict) else None
                        if cat and str(value).strip() != "":
                            parts.append(f"向{cat}输入框填写{value}")
                            continue  # 已输出定制语句
                    except Exception:
                        pass
            meta = mapping.get(norm_path)
            if not meta:
                continue  # 未登记的字段不参与操作
            mark = (meta.get("type") or "").lower().strip()
            label = meta.get("label") or norm_path  # 无中文名则退回路径
            # 空标记：参考信息，跳过
            if mark == "":
                continue

            if mark == "i":
                parts.append(f"在{label}输入框中输入{value}")
            elif mark == "c":
                parts.append(f"在{label}下拉框中选择值为\"{value}\"")
            elif mark == "r":
                parts.append(f"点击{label}radio button")
            elif mark == "d":
                parts.append(f"选择日期{label}为{value}")
            elif mark == "empty":
                parts.append(f"{label}内容为{value}")
            else:
                # 未知标记，按参考信息忽略
                continue

        return "。".join(parts) + ("。" if parts else "")

    def _build_prompt_reimbursement_flow(self, data: Dict[str, Any], mapping: Dict[str, Dict[str, str]]) -> str:
        """按报销业务的阶段顺序构建提示词，并在每一阶段后拼接跳转说明。"""
        stage_order = [
            ("login", self.LOGIN_TRANSITION),
            ("project", self.PROJECT_TRANSITION),
            ("expenses", self.EXPENSE_TRANSITION),
            ("personnel", self.PERSONNEL_TRANSITION),
            ("appointment", self.APPOINTMENT_TRANSITION),
        ]
        segments: List[str] = []
        # 添加打开网页指令
        segments.append("请你调用Playwright MCP，执行以下命令，一次性执行完")
        segments.append("打开https://cwcx.uestc.edu.cn/WFManager/login.jsp")
        # 开头标题
        segments.append("业务大类：报销业务。以下是需要执行的页面操作：")

        for stage_key, transition_text in stage_order:
            actions = self._generate_actions_for_stage(data, stage_key, mapping)
            if actions:
                # 阶段动作句子，每个动作换行
                for i, action in enumerate(actions):
                    segments.append(action)
            # 在 laborInfo 阶段填写完毕后，插入跳转到 laborPerson 的说明
            if stage_key == "laborInfo":
                segments.extend([line.strip() for line in self.LABORINFO_TO_LABORPERSON_TRANSITION.split('\n') if line.strip()])
            # 阶段跳转说明（无论是否有动作，都附加，便于保持固定流程）
            if transition_text:
                # 将跳转文本按行分割，每行单独作为一个segment
                transition_lines = [line.strip() for line in transition_text.split('\n') if line.strip()]
                segments.extend(transition_lines)
        
        # 为每行添加序号
        numbered_segments = []
        for i, segment in enumerate(segments, 1):
            if segment.strip():
                numbered_segments.append(f"{i}. {segment.strip()}")
        
        return "\n".join(numbered_segments)

    def _build_prompt_travel_flow(self, data: Dict[str, Any], mapping: Dict[str, Dict[str, str]]) -> str:
        """按业务出差旅费的阶段顺序构建提示词，并在每一阶段后拼接跳转说明。"""
        stage_order = [
            ("login", self.LOGIN_TRANSITION),
            ("project", self.PROJECT_TRANSITION),
            ("travelPerson", ""),  # 不附加跳转说明
            ("travelExpenses", self.TRAVEL_TRANSITION),  # 只在travelExpenses后附加
            ("personnel", self.PERSONNEL_TRANSITION),
            ("appointment", self.APPOINTMENT_TRANSITION),
        ]
        segments: List[str] = []
        # 添加打开网页指令
        segments.append("请你调用Playwright MCP，执行以下命令，一次性执行完")
        segments.append("打开https://cwcx.uestc.edu.cn/WFManager/login.jsp")
        # 开头标题
        segments.append("业务大类：业务出差旅费。以下是需要执行的页面操作：")

        for stage_key, transition_text in stage_order:
            actions = self._generate_actions_for_stage(data, stage_key, mapping)
            if actions:
                # 阶段动作句子，每个动作换行
                for i, action in enumerate(actions):
                    segments.append(action)
            # 阶段跳转说明（无论是否有动作，都附加，便于保持固定流程）
            if transition_text:
                # 将跳转文本按行分割，每行单独作为一个segment
                transition_lines = [line.strip() for line in transition_text.split('\n') if line.strip()]
                segments.extend(transition_lines)
        
        # 为每行添加序号
        numbered_segments = []
        for i, segment in enumerate(segments, 1):
            if segment.strip():
                numbered_segments.append(f"{i}. {segment.strip()}")
        
        return "\n".join(numbered_segments)

    def _build_prompt_labor_flow(self, data: Dict[str, Any], mapping: Dict[str, Dict[str, str]]) -> str:
        """按酬金业务的阶段顺序构建提示词，并在每一阶段后拼接跳转说明。"""
        stage_order = [
            ("login", self.LOGIN_TRANSITION),
            ("project", self.PROJECT_TRANSITION),
            ("laborInfo", self.LABOR_TRANSITION),
            ("laborPerson", self.PERSONNEL_TRANSITION),
            ("appointment", self.APPOINTMENT_TRANSITION),
        ]
        segments: List[str] = []
        # 添加打开网页指令
        segments.append("请你调用Playwright MCP，执行以下命令，一次性执行完")
        segments.append("打开https://cwcx.uestc.edu.cn/WFManager/login.jsp")
        # 开头标题
        segments.append("业务大类：酬金业务。以下是需要执行的页面操作：")

        for stage_key, transition_text in stage_order:
            actions = self._generate_actions_for_stage(data, stage_key, mapping)
            if actions:
                # 阶段动作句子，每个动作换行
                for i, action in enumerate(actions):
                    segments.append(action)
            # 阶段跳转说明（无论是否有动作，都附加，便于保持固定流程）
            if transition_text:
                # 将跳转文本按行分割，每行单独作为一个segment
                transition_lines = [line.strip() for line in transition_text.split('\n') if line.strip()]
                segments.extend(transition_lines)
        
        # 为每行添加序号
        numbered_segments = []
        for i, segment in enumerate(segments, 1):
            if segment.strip():
                numbered_segments.append(f"{i}. {segment.strip()}")
        
        return "\n".join(numbered_segments)

    def _generate_actions_for_stage(self, data: Dict[str, Any], stage_key: str, mapping: Dict[str, Dict[str, str]]) -> List[str]:
        """针对某个顶层阶段键（如 'login'、'project'、'expenses'），
        从数据中筛选出该阶段下的所有字段，依据Excel映射生成动作语句。"""
        if stage_key not in data or data.get(stage_key) in (None, ""):
            return []
        stage_actions: List[str] = []

        # 特殊处理：expenses阶段的amount，使用对应category作为控件名
        if stage_key == "expenses" and isinstance(data.get("expenses"), list):
            for item in data.get("expenses") or []:
                if not isinstance(item, dict):
                    continue
                cat = item.get("category")
                amt = item.get("amount")
                if cat is not None and amt not in (None, ""):
                    stage_actions.append(f"向{cat}输入框填写{amt}")
        
        # 特殊处理：travelPerson阶段的多个出差人员
        elif stage_key == "travelPerson" and isinstance(data.get("travelPerson"), list):
            travel_person_list = data.get("travelPerson") or []
            print(f"DEBUG: 处理travelPerson阶段，共{len(travel_person_list)}个出差人员")
            for person_idx, travel_person in enumerate(travel_person_list):
                if not isinstance(travel_person, dict):
                    continue
                
                print(f"DEBUG: 处理第{person_idx}个出差人员: {travel_person}")
                
                # 生成该出差人员的所有字段动作
                person_actions = []
                for path, value in self._flatten_json(travel_person):
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        continue
                    
                    # 构建完整路径
                    full_path = f"travelPerson[{person_idx}].{path}" if path != "travelPerson" else f"travelPerson[{person_idx}]"
                    norm_path = self._normalize_array_path(full_path)
                    
                    print(f"DEBUG: 字段路径: {path} -> {full_path} -> {norm_path}, 值: {value}")
                    
                    meta = mapping.get(norm_path)
                    print(f"DEBUG: 映射信息: {meta}")
                    
                    if not meta:
                        print(f"DEBUG: 未找到映射，跳过字段: {norm_path}")
                        continue
                    
                    mark = (meta.get("type") or "").lower().strip()
                    label = meta.get("label") or norm_path
                    
                    print(f"DEBUG: 标记: {mark}, 标签: {label}")
                    
                    if mark == "":
                        print(f"DEBUG: 标记为空，跳过字段: {norm_path}")
                        continue
                    
                    if mark == "i":
                        action = f"在{label}输入框中输入{value}"
                        person_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "c":
                        action = f"在{label}下拉框中选择值为\"{value}\""
                        person_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "r":
                        action = f"点击{label}radio button"
                        person_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "d":
                        action = f"选择日期{label}为{value}"
                        person_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "empty":
                        action = f"{label}内容为{value}"
                        person_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                
                # 添加该出差人员的动作
                stage_actions.extend(person_actions)
                print(f"DEBUG: 当前出差人员动作数: {len(person_actions)}")
            
            print(f"DEBUG: travelPerson阶段总动作数: {len(stage_actions)}")
            return stage_actions
        
        # 特殊处理：travelExpenses阶段的费用信息
        elif stage_key == "travelExpenses" and isinstance(data.get("travelExpenses"), list):
            travel_expenses_list = data.get("travelExpenses") or []
            print(f"DEBUG: 处理travelExpenses阶段，共{len(travel_expenses_list)}个费用项目")
            for expense_idx, travel_expense in enumerate(travel_expenses_list):
                if not isinstance(travel_expense, dict):
                    continue
                
                print(f"DEBUG: 处理第{expense_idx}个费用项目: {travel_expense}")
                
                # 生成该费用项目的所有字段动作
                expense_actions = []
                for path, value in self._flatten_json(travel_expense):
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        continue
                    
                    # 构建完整路径
                    full_path = f"travelExpenses[{expense_idx}].{path}" if path != "travelExpenses" else f"travelExpenses[{expense_idx}]"
                    norm_path = self._normalize_array_path(full_path)
                    
                    print(f"DEBUG: 字段路径: {path} -> {full_path} -> {norm_path}, 值: {value}")
                    
                    meta = mapping.get(norm_path)
                    print(f"DEBUG: 映射信息: {meta}")
                    
                    if not meta:
                        print(f"DEBUG: 未找到映射，跳过字段: {norm_path}")
                        continue
                    
                    mark = (meta.get("type") or "").lower().strip()
                    label = meta.get("label") or norm_path
                    
                    print(f"DEBUG: 标记: {mark}, 标签: {label}")
                    
                    if mark == "":
                        print(f"DEBUG: 标记为空，跳过字段: {norm_path}")
                        continue
                    
                    if mark == "i":
                        action = f"在{label}输入框中输入{value}"
                        expense_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "c":
                        action = f"在{label}下拉框中选择值为\"{value}\""
                        expense_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "r":
                        action = f"点击{label}radio button"
                        expense_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "d":
                        action = f"选择日期{label}为{value}"
                        expense_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                    elif mark == "empty":
                        action = f"{label}内容为{value}"
                        expense_actions.append(action)
                        print(f"DEBUG: 添加动作: {action}")
                
                # 添加该费用项目的动作
                stage_actions.extend(expense_actions)
                print(f"DEBUG: 当前费用项目动作数: {len(expense_actions)}")
            
            print(f"DEBUG: travelExpenses阶段总动作数: {len(stage_actions)}")
            return stage_actions
        
        # 特殊处理：laborInfo阶段的劳务信息（单对象）
        elif stage_key == "laborInfo" and isinstance(data.get("laborInfo"), dict):
            labor_item = data.get("laborInfo")
            if not isinstance(labor_item, dict):
                return []
            
            # 生成该劳务信息的字段动作（只处理laborType和reason字段）
            labor_actions = []
            for path, value in self._flatten_json(labor_item):
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    continue
                
                # 只处理laborInfo相关字段
                if path not in ["laborType", "reason", "personnelCategory", "remunerationNature", "remarks", "paymentStandard", "startTime", "endTime"]:
                    continue
                
                # 构建完整路径（与Excel映射一致）
                full_path = f"laborInfo.{path}"
                norm_path = self._normalize_array_path(full_path)
                
                meta = mapping.get(norm_path)
                if not meta:
                    continue
                mark = (meta.get("type") or "").lower().strip()
                label = meta.get("label") or norm_path
                if mark == "":
                    continue
                
                if mark == "i":
                    labor_actions.append(f"在{label}输入框中输入{value}")
                elif mark == "c":
                    labor_actions.append(f"在{label}下拉框中选择值为\"{value}\"")
                elif mark == "r":
                    labor_actions.append(f"点击{label}radio button")
                elif mark == "d":
                    labor_actions.append(f"选择日期{label}为{value}")
                elif mark == "empty":
                    labor_actions.append(f"{label}内容为{value}")
            
            # 添加该劳务信息的动作
            stage_actions.extend(labor_actions)
            
            return stage_actions
        
        # 特殊处理：laborPerson阶段的多个劳务人员，每个人后添加LABOR_ITEM_POST_ACTION
        elif stage_key == "laborPerson" and isinstance(data.get("laborPerson"), list):
            labor_list = data.get("laborPerson") or []
            for labor_idx, labor_person in enumerate(labor_list):
                if not isinstance(labor_person, dict):
                    continue
                
                # 生成该劳务人员的字段动作（只处理amount字段）
                labor_actions = []
                # 在每位人员信息开始前，插入预操作说明
                pre_lines = [line.strip() for line in self.LABORPERSON_PRE_ACTION.split('\n') if line.strip()]
                stage_actions.extend(pre_lines)
                for path, value in self._flatten_json(labor_person):
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        continue
                    
                    # 只处理laborPerson相关字段
                    if path not in ["employeeId", "amount", "singleEntryAmount"]:
                        continue
                    
                    # 构建完整路径（与Excel映射一致）
                    full_path = f"laborPerson[{labor_idx}].{path}"
                    norm_path = self._normalize_array_path(full_path)
                    
                    meta = mapping.get(norm_path)
                    if not meta:
                        # 映射缺失时的容错：提供默认类型与中文名称
                        default_labels = {
                            "laborPerson[].employeeId": "工号/证件号",
                            "laborPerson[].amount": "金额",
                            "laborPerson[].singleEntryAmount": "单笔录入金额",
                        }
                        meta = {"type": "i", "label": default_labels.get(norm_path, norm_path)}
                    mark = (meta.get("type") or "").lower().strip()
                    label = meta.get("label") or norm_path
                    if mark == "":
                        continue
                    
                    if mark == "i":
                        labor_actions.append(f"在{label}输入框中输入{value}")
                    elif mark == "c":
                        labor_actions.append(f"在{label}下拉框中选择值为\"{value}\"")
                    elif mark == "r":
                        labor_actions.append(f"点击{label}radio button")
                    elif mark == "d":
                        labor_actions.append(f"选择日期{label}为{value}")
                    elif mark == "empty":
                        labor_actions.append(f"{label}内容为{value}")
                
                # 添加该劳务人员的动作
                stage_actions.extend(labor_actions)
                
                # 如果不是最后一个劳务人员，添加LABOR_ITEM_POST_ACTION
                if labor_idx < len(labor_list) - 1:
                    # 将LABOR_ITEM_POST_ACTION按行分割，每行单独添加
                    post_action_lines = [line.strip() for line in self.LABOR_ITEM_POST_ACTION.split('\n') if line.strip()]
                    stage_actions.extend(post_action_lines)
            
            return stage_actions
        
        # 特殊处理：personnel阶段的多个人员，每个人后添加PERSONNEL_ITEM_POST_ACTION
        elif stage_key == "personnel" and isinstance(data.get("personnel"), list):
            personnel_list = data.get("personnel") or []
            for person_idx, person in enumerate(personnel_list):
                if not isinstance(person, dict):
                    continue
                
                # 生成该人员的所有字段动作
                person_actions = []
                for path, value in self._flatten_json(person):
                    if value is None or (isinstance(value, str) and value.strip() == ""):
                        continue
                    
                    # 构建完整路径
                    full_path = f"personnel[{person_idx}].{path}" if path != "personnel" else f"personnel[{person_idx}]"
                    norm_path = self._normalize_array_path(full_path)
                    
                    meta = mapping.get(norm_path)
                    if not meta:
                        continue
                    mark = (meta.get("type") or "").lower().strip()
                    label = meta.get("label") or norm_path
                    if mark == "":
                        continue
                    
                    if mark == "i":
                        person_actions.append(f"在{label}输入框中输入{value}")
                    elif mark == "c":
                        person_actions.append(f"在{label}下拉框中选择值为\"{value}\"")
                    elif mark == "r":
                        person_actions.append(f"点击{label}radio button")
                    elif mark == "d":
                        person_actions.append(f"选择日期{label}为{value}")
                    elif mark == "empty":
                        person_actions.append(f"{label}内容为{value}")
                
                # 添加该人员的动作
                stage_actions.extend(person_actions)
                
                # 如果不是最后一个人员，添加PERSONNEL_ITEM_POST_ACTION
                if person_idx < len(personnel_list) - 1:
                    # 将PERSONNEL_ITEM_POST_ACTION按行分割，每行单独添加
                    post_action_lines = [line.strip() for line in self.PERSONNEL_ITEM_POST_ACTION.split('\n') if line.strip()]
                    stage_actions.extend(post_action_lines)
            
            return stage_actions
        
        # 普通处理：遍历所有 (path, value) 对，筛选以 stage_key 开头的路径
        for path, value in self._flatten_json(data):
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            # 只处理该阶段路径
            if not (path == stage_key or path.startswith(stage_key + ".") or path.startswith(stage_key + "[")):
                continue
            # 已经为expenses[].amount生成过定制语句，避免重复
            norm_path = self._normalize_array_path(path)
            if stage_key == "expenses" and norm_path == "expenses[].amount":
                continue
            # 已经为travelPerson和travelExpenses处理过，避免重复
            if stage_key in ["travelPerson", "travelExpenses"]:
                continue
            # 已经为laborInfo和laborPerson处理过，避免重复
            if stage_key in ["laborInfo", "laborPerson"]:
                continue
            # 已经为personnel处理过，避免重复
            if stage_key == "personnel":
                continue

            meta = mapping.get(norm_path)
            if not meta:
                continue
            mark = (meta.get("type") or "").lower().strip()
            label = meta.get("label") or norm_path
            if mark == "":
                continue
            if mark == "i":
                stage_actions.append(f"在{label}输入框中输入{value}")
            elif mark == "c":
                stage_actions.append(f"在{label}下拉框中选择值为\"{value}\"")
            elif mark == "r":
                stage_actions.append(f"点击{label}radio button")
            elif mark == "d":
                stage_actions.append(f"选择日期{label}为{value}")
            elif mark == "empty":
                stage_actions.append(f"{label}内容为{value}")
        return stage_actions

    def build_stage_prompts_from_data(self, data: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
        """根据JSON生成按阶段划分的Playwright提示词。"""
        mapping = self.load_field_type_mapping()
        business_type = (data.get("businessType") or "").strip()
        stage_flow = self._get_stage_flow_config(business_type)
        if not stage_flow:
            return {}

        stage_prompts: Dict[str, Dict[str, str]] = {}
        for idx, stage in enumerate(stage_flow):
            stage_key = stage.get("key")
            stage_lines: List[str] = []

            if idx == 0:
                stage_lines.append("请你调用Playwright MCP，执行以下命令，一次性执行完")
                stage_lines.append("打开https://cwcx.uestc.edu.cn/WFManager/login.jsp")
                if business_type:
                    stage_lines.append(f"业务大类：{business_type}。以下是需要执行的页面操作：")

            actions = self._generate_actions_for_stage(data, stage_key, mapping)
            stage_lines.extend(actions)

            if stage_key == "laborInfo":
                stage_lines.extend(
                    [line.strip() for line in self.LABORINFO_TO_LABORPERSON_TRANSITION.split('\n') if line.strip()]
                )

            transition_text = stage.get("transition") or ""
            if transition_text:
                transition_lines = [line.strip() for line in transition_text.split('\n') if line.strip()]
                stage_lines.extend(transition_lines)

            stage_lines = [line.strip() for line in stage_lines if line and line.strip()]
            if not stage_lines:
                continue

            stage_prompts[stage_key] = {
                "title": stage.get("title") or stage_key,
                "prompt": self._format_segments(stage_lines)
            }

        return stage_prompts

    def build_playwright_prompt_from_input(self, user_input: str) -> Dict[str, Any]:
        """一体化：提取→读取映射→生成提示词。

        Returns:
            { prompt: str, data: dict }
        """
        data = self.extract_form_json(user_input)
        if not isinstance(data, dict) or "error" in data:
            return {"error": data.get("error", "提取失败"), "data": data}

        # 只读取现有映射，不修改Excel文件
        prompt = self.build_playwright_prompt_from_data(data)
        return {"prompt": prompt, "data": data}

    # 新增：API风格方法，入参为自然语言，直接返回MCP提示词字符串
    def generate_mcp_prompt_from_nl(self, nl_text: str) -> str:
        """将报销自然语言直接转换为 MCP 提示词字符串。

        Args:
            nl_text: 报销相关自然语言

        Returns:
            仅返回生成的 MCP 提示词字符串；若失败则返回空字符串
        """
        result = self.build_playwright_prompt_from_input(nl_text)
        if isinstance(result, dict):
            if result.get("error"):
                return ""
            return result.get("prompt", "")
        return result if isinstance(result, str) else ""

if __name__ == "__main__":
    workflow = WorkflowCore()
    try:
        print("=== 报销自动化 · 交互式生成MCP提示词 ===")
        print("提示：Excel映射表路径为 field_type_mapping.xlsx（与本文件同目录）\n")
        user_text = input("请输入报销相关自然语言（直接回车结束）：\n> ").strip()
        if not user_text:
            print("未输入内容，已退出。")
        else:
            result = workflow.build_playwright_prompt_from_input(user_text)
            if isinstance(result, dict) and result.get("error"):
                print(f"生成失败：{result.get('error')}")
                if result.get("data"):
                    print(result.get("data"))
            else:
                prompt = result.get("prompt", "") if isinstance(result, dict) else ""
                data = result.get("data", {}) if isinstance(result, dict) else {}
                try:
                    print("\n=== 提取到的JSON数据 ===")
                    print(json.dumps(data, ensure_ascii=False, indent=2))
                except Exception:
                    print(data)
                print("\n=== 生成的Playwright MCP提示词 ===")
                print(prompt or "(无可用指令，请检查Excel映射或输入内容)")
    except KeyboardInterrupt:
        print("\n已取消。")


def process_excel_to_mcp_prompt(excel_path: str, sheet_name: str = None) -> List[str]:
    """
    从Excel文件生成自然语言，然后转换为MCP提示词
    
    Args:
        excel_path: Excel文件路径
        sheet_name: 工作表名称（可选）
        
    Returns:
        每个序号的MCP提示词列表
    """
    try:
        # 导入Excel处理模块
        from excel_to_nl import generate_nl_from_excel
        
        # 从Excel生成自然语言
        nl_summaries = generate_nl_from_excel(excel_path, sheet_name)
        
        # 初始化工作流程控制器
        workflow = WorkflowCore()
        
        # 将每个自然语言转换为MCP提示词
        mcp_prompts = []
        for i, nl_text in enumerate(nl_summaries):
            print(f"=== 处理序号 {i+1} ===")
            print(f"自然语言: {nl_text}")
            
            # 提取JSON并生成MCP提示词
            json_data = workflow.extract_form_json(nl_text)
            if json_data:
                mcp_prompt = workflow.build_playwright_prompt_from_data(json_data)
                mcp_prompts.append(mcp_prompt)
                print(f"MCP提示词已生成")
            else:
                print(f"无法从序号 {i+1} 中提取有效信息")
                mcp_prompts.append("")
            print("-" * 50)
        
        return mcp_prompts
        
    except Exception as e:
        print(f"处理Excel文件时发生错误: {e}")
        return []


def process_excel_to_mcp_direct(excel_path: str, sheet_name: str, serial) -> str:
    """
    直接从Excel生成MCP提示词（跳过LLM自然语言生成环节）
    
    Args:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        serial: 目标序号（字符串或整数）
        
    Returns:
        生成的MCP提示词字符串（失败返回空字符串）
    """
    try:
        # 导入Excel处理模块
        from excel_to_nl import excel_to_json_direct
        
        # 直接从Excel转JSON
        json_data = excel_to_json_direct(excel_path, sheet_name, serial)
        
        if not json_data:
            print(f"未找到序号 {serial} 的数据")
            return ""
        
        # 初始化工作流程控制器并生成MCP提示词
        workflow = WorkflowCore()
        mcp_prompt = workflow.build_playwright_prompt_from_data(json_data)
        
        return mcp_prompt
        
    except Exception as e:
        print(f"处理Excel文件时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return ""


def process_excel_to_stage_prompts(excel_path: str, sheet_name: str, serial) -> Dict[str, Any]:
    """
    从 Excel 生成按阶段划分的 MCP 提示词。
    """
    try:
        from excel_to_nl import excel_to_json_direct

        json_data = excel_to_json_direct(excel_path, sheet_name, serial)
        if not json_data:
            print(f"未找到序号 {serial} 的数据")
            return {}

        workflow = WorkflowCore()
        full_prompt = workflow.build_playwright_prompt_from_data(json_data)
        stage_prompts = workflow.build_stage_prompts_from_data(json_data)

        return {
            "full_prompt": full_prompt,
            "stage_prompts": stage_prompts,
            "business_type": json_data.get("businessType"),
            "serial": str(serial),
        }
    except Exception as e:
        print(f"生成阶段提示词时发生错误: {e}")
        traceback.print_exc()
        return {}


def batch_process_excel_to_mcp_direct(excel_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    """
    批量处理Excel所有序号，直接生成MCP提示词（跳过LLM）
    只处理"!已生成MCP提示词"列为空的序号
    
    Args:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        
    Returns:
        包含每个序号的JSON数据和MCP提示词的列表
    """
    try:
        # 导入Excel处理模块
        from excel_to_nl import excel_to_json_direct
        from openpyxl import load_workbook
        
        # 读取所有序号
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # 查找"!已生成MCP提示词"列的索引
        headers = {}
        for idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip()] = idx
        
        mcp_col_idx = headers.get("!已生成MCP提示词")
        serial_col_idx = headers.get("序号", 1)  # 默认第1列是序号
        
        # 获取所有未处理的序号（"!已生成MCP提示词"列为空）
        serials_to_process = []
        skipped_serials = []
        
        for row in ws.iter_rows(min_row=2):
            serial_cell = row[serial_col_idx - 1]  # 列索引转为0-based
            if serial_cell.value:
                serial_val = str(serial_cell.value).strip()
                
                # 检查"!已生成MCP提示词"列
                if mcp_col_idx:
                    mcp_cell = row[mcp_col_idx - 1]
                    if mcp_cell.value and str(mcp_cell.value).strip():
                        # 已有值，跳过
                        skipped_serials.append(serial_val)
                        continue
                
                # 未处理，添加到待处理列表
                serials_to_process.append(serial_val)
        
        # 去重并排序
        serials_to_process = sorted(set(serials_to_process), key=lambda x: int(x) if x.isdigit() else 0)
        
        if skipped_serials:
            print(f"ℹ️  跳过已生成的序号: {', '.join(sorted(set(skipped_serials), key=lambda x: int(x) if x.isdigit() else 0))}")
        
        if not serials_to_process:
            print("ℹ️  所有序号都已生成MCP提示词，无需处理")
            return []
        
        print(f"ℹ️  待处理序号: {', '.join(serials_to_process)}")
        print()
        
        # 初始化工作流程控制器
        workflow = WorkflowCore()
        
        # 批量处理
        results = []
        for serial in serials_to_process:
            print(f"=== 处理序号 {serial} ===")
            
            # 直接从Excel转JSON
            json_data = excel_to_json_direct(excel_path, sheet_name, serial)
            
            if not json_data:
                print(f"序号 {serial}: 未找到数据")
                continue
            
            # 生成MCP提示词
            mcp_prompt = workflow.build_playwright_prompt_from_data(json_data)
            
            results.append({
                "serial": serial,
                "json_data": json_data,
                "mcp_prompt": mcp_prompt
            })
            
            print(f"序号 {serial}: MCP提示词已生成（{len(mcp_prompt)} 字符）")
        
        return results
        
    except Exception as e:
        print(f"批量处理Excel文件时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return []


def mark_mcp_file_as_completed(excel_path: str, sheet_name: str, old_filename: str) -> str:
    """
    将MCP提示词文件标记为已完成（未预约→已预约），并更新Excel
    
    Args:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        old_filename: 原文件名（未预约-xxx.txt）
        
    Returns:
        新文件名（已预约-xxx.txt）
    """
    try:
        from openpyxl import load_workbook
        import os
        
        # 构建完整路径
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, "mcp_prompts")
        old_filepath = os.path.join(output_dir, old_filename)
        
        # 生成新文件名
        new_filename = old_filename.replace("未预约-", "已预约-")
        new_filepath = os.path.join(output_dir, new_filename)
        
        # 重命名文件
        if os.path.exists(old_filepath):
            os.rename(old_filepath, new_filepath)
            print(f"✅ 文件已重命名: {old_filename} → {new_filename}")
        else:
            print(f"⚠️  文件不存在: {old_filename}")
            return old_filename
        
        # 更新Excel中的记录
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]
        
        # 查找"!已生成MCP提示词"列
        headers = {}
        for idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip()] = idx
        
        mcp_col_idx = headers.get("!已生成MCP提示词")
        
        if mcp_col_idx:
            # 查找包含旧文件名的单元格并更新
            for row_idx in range(2, ws.max_row + 1):
                mcp_cell = ws.cell(row=row_idx, column=mcp_col_idx)
                if mcp_cell.value and old_filename in str(mcp_cell.value):
                    mcp_cell.value = new_filename
            
            wb.save(excel_path)
            print(f"✅ Excel文件已更新")
        
        wb.close()
        return new_filename
        
    except Exception as e:
        print(f"⚠️  标记完成时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return old_filename
