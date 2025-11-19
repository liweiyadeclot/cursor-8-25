#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
驱动脚本：遍历读取 420财务050823.xlsx 的 `3-报销` 表中每一个序号，
依次执行：excel_to_nl.generate_single_nl_from_excel → WorkflowCore.generate_mcp_prompt_from_nl，
得到 MCP 提示词后，尝试通过 Playwright MCP 执行。

注意：
- 需要本地 Qwen（Ollama）可用以提升 NL 质量；不可用则退回模板。
- 需要已有 `field_type_mapping.xlsx` 来指导控件类型解析；若缺失则提示并继续。
- Playwright MCP 通过 HTTP 接口调用，如无可用端点则仅打印提示词。
"""

from __future__ import annotations

import os
import sys
import json
import time
from typing import Optional

# 将当前脚本所在目录加入路径，确保本地相对导入
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
if CURRENT_DIR not in sys.path:
    sys.path.insert(0, CURRENT_DIR)

try:
    from excel_to_nl import generate_single_nl_from_excel
except Exception as e:
    print(f"导入 excel_to_nl 失败: {e}")
    raise

try:
    from workflow_core import WorkflowCore
except Exception as e:
    print(f"导入 workflow_core 失败: {e}")
    raise


def call_playwright_mcp(prompt: str) -> bool:
    """尝试调用 Playwright MCP 执行指令。

    这里提供两种常见方式：
    1) HTTP: 通过本地 MCP 网关（若有）转发；
    2) STDIO/CLI: 若没有 HTTP 服务，则仅打印提示词供人工/其他进程使用。

    返回是否成功调用的布尔值（仅代表请求是否发送，不代表网页自动化成功）。
    """
    prompt = (prompt or "").strip()
    if not prompt:
        print("[MCP] 提示词为空，跳过调用。")
        return False

    # 优先尝试 HTTP 接口（若用户已部署 MCP 网关）
    mcp_http = os.environ.get("MCP_HTTP_ENDPOINT")
    if mcp_http:
        try:
            import requests
            resp = requests.post(mcp_http, json={"prompt": prompt}, timeout=30)
            print(f"[MCP] HTTP 调用状态: {resp.status_code}")
            if resp.status_code == 200:
                try:
                    print("[MCP] 返回:")
                    print(resp.json())
                except Exception:
                    print(resp.text)
                return True
        except Exception as e:
            print(f"[MCP] HTTP 调用失败: {e}")

    # 退回：打印提示词，由外部 Playwright MCP 客户端处理
    print("\n================ MCP 提示词（请交由 Playwright MCP 执行） ================")
    print(prompt)
    print("====================================================================\n")
    return False


def main(
    excel_path: Optional[str] = None,
    sheet_name: str = "3-报销",
    use_llm: bool = True,
) -> int:
    # 默认 Excel 路径（仓库内）
    if not excel_path:
        excel_path = os.path.join(os.path.dirname(CURRENT_DIR), "420财务050823.xlsx")

    if not os.path.exists(excel_path):
        print(f"错误：Excel 文件不存在：{excel_path}")
        return 1

    print(f"读取 Excel：{excel_path}，工作表：{sheet_name}")

    # 预扫描序号范围：通过 openpyxl 简单遍历第一列“序号”（或按 excel_to_nl 的分组逻辑也可）
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print("请先安装依赖：pip install openpyxl")
        print(e)
        return 2

    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 读取表头映射
    headers = {}
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, val in enumerate(header_row):
        key = str(val or "").strip()
        if key:
            headers[key] = idx

    # 收集数据行与按 excel_to_nl 的逻辑逐序号递增
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

    # 构造 groups 逻辑与 excel_to_nl 保持一致（遇到空序号延续上一个）
    from collections import defaultdict
    groups = defaultdict(list)
    current_serial = "1"
    def _get(row_tuple, name: str, default: str = "") -> str:
        idx = headers.get(name)
        if idx is None:
            return default
        v = row_tuple[idx]
        return str(v).strip() if v is not None else default

    for r in data_rows:
        serial = _get(r, "序号") or _get(r, "编号")
        if serial:
            current_serial = str(serial)
        groups[current_serial].append(r)

    # 逐序号处理
    workflow = WorkflowCore()
    all_serials = sorted(groups.keys(), key=lambda x: str(x))
    print(f"共发现序号组：{len(all_serials)} 个 -> {all_serials}")

    for serial in all_serials:
        print("-" * 80)
        print(f"开始处理 序号={serial}")
        try:
            nl_text = generate_single_nl_from_excel(
                filepath=excel_path,
                sheet_name=sheet_name,
                serial=serial,
                use_llm=use_llm,
            )
        except Exception as e:
            print(f"生成自然语言失败（序号={serial}）：{e}")
            continue

        nl_text = (nl_text or "").strip()
        if not nl_text:
            print(f"序号={serial} 未得到自然语言，跳过。")
            continue

        print("[NL] 自然语言：")
        print(nl_text)

        try:
            mcp_prompt = workflow.generate_mcp_prompt_from_nl(nl_text)
        except Exception as e:
            print(f"生成 MCP 提示词失败（序号={serial}）：{e}")
            continue

        if not mcp_prompt:
            print(f"序号={serial} 未生成有效的 MCP 提示词，可能缺少映射或提取失败。")
            continue

        print("[MCP] 生成的提示词：")
        print(mcp_prompt)

        # 调用 Playwright MCP 执行
        _ = call_playwright_mcp(mcp_prompt)

        # 防抖间隔，避免过快连续触发
        time.sleep(1)

    print("已处理所有序号。")
    return 0


if __name__ == "__main__":
    sys.exit(main())





