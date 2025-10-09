#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将Excel中的行记录（按“序号”分组）转换为总结性的报销自然语言。

用法（命令行）：
  python excel_to_nl.py path/to/file.xlsx [sheet_name]

返回：按序号分组后的自然语言列表，逐条打印。
"""

from __future__ import annotations

import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional
import os
import json
import requests

try:
    from openpyxl import load_workbook
except Exception as e:  # pragma: no cover
    raise RuntimeError("请先安装依赖：pip install openpyxl") from e


HeaderMap = Dict[str, int]


def _read_sheet_headers(ws) -> HeaderMap:
    """读取第一行作为表头，返回 列名->列索引 的映射（0-based）。"""
    headers: HeaderMap = {}
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, val in enumerate(row):
        key = str(val or "").strip()
        if key:
            headers[key] = idx
    return headers


def _get(row: tuple, headers: HeaderMap, name: str, default: str = "") -> str:
    idx = headers.get(name)
    if idx is None:
        return default
    val = row[idx]
    if val is None:
        return default
    return str(val).strip()


def _group_rows_by_serial(rows: List[tuple], headers: HeaderMap) -> Dict[str, List[tuple]]:
    """按"序号"分组。"""
    groups: Dict[str, List[tuple]] = defaultdict(list)
    current_serial = "1"  # 默认序号
    
    for r in rows:
        # 如果当前行有序号，更新当前序号
        serial = _get(r, headers, "序号") or _get(r, headers, "编号")
        if serial:
            current_serial = serial
        
        # 将当前行添加到当前序号组
        groups[current_serial].append(r)
    
    return groups


def _normalize_number(text: str) -> str:
    """提取数字（若包含单位则去除），保持原值为数字字符串。"""
    import re
    m = re.search(r"[-+]?[0-9]*\.?[0-9]+", text or "")
    return m.group(0) if m else text


def _build_summary_for_group(rows: List[tuple], headers: HeaderMap) -> str:
    """将同一“序号”下的多行合成为一段自然语言。"""
    # 公共字段（取第一行）
    first = rows[0]
    project_no = _get(first, headers, "报销项目号") or _get(first, headers, "项目号")
    attach_cnt = _normalize_number(_get(first, headers, "附件张数") or _get(first, headers, "附件"))
    pay_method = _get(first, headers, "支付方式")
    remark = _get(first, headers, "备注")
    special = _get(first, headers, "特殊事项说明")
    date = _get(first, headers, "日期")
    campus = _get(first, headers, "校区")

    # 转卡信息（汇总所有行）
    transfer_info_items: List[str] = []
    for r in rows:
        emp_id = _get(r, headers, "转卡信息工号") or _get(r, headers, "工号")
        card_tail = _get(r, headers, "卡号尾号").lstrip("*")
        personal_amount = _normalize_number(_get(r, headers, "个人金额"))
        
        if emp_id or card_tail or personal_amount:
            transfer_parts = []
            if emp_id:
                transfer_parts.append(f"工号{emp_id}")
            if card_tail:
                transfer_parts.append(f"卡号尾号{card_tail}")
            if personal_amount:
                transfer_parts.append(f"个人金额{personal_amount}")
            if transfer_parts:
                transfer_info_items.append("，".join(transfer_parts))

    # 科目信息（汇总所有行）
    expense_items: List[str] = []
    for r in rows:
        subject = _get(r, headers, "科目")
        amt = _normalize_number(_get(r, headers, "金额"))
        if subject or amt:
            if amt:
                expense_items.append(f"{subject}{amt}")
            else:
                expense_items.append(subject)

    expense_text = "、".join(expense_items) if expense_items else "无科目"

    # 业务大类推断：这里默认报销业务
    business_type = "报销业务"

    # 组织自然语言
    parts: List[str] = []
    parts.append(f"业务大类为{business_type}")
    if project_no:
        parts.append(f"报销项目号{project_no}")
    if attach_cnt:
        parts.append(f"附件张数为{attach_cnt}张")
    if pay_method:
        parts.append(f"支付方式为{pay_method}")
    if remark:
        parts.append(f"备注“{remark}”")
    if special:
        parts.append(f"特殊事项说明“{special}”")
    if expense_text:
        parts.append(f"科目与金额包括：{expense_text}")
    if transfer_info_items:
        parts.append(f"转卡信息：{';'.join(transfer_info_items)}")
    if date:
        parts.append(f"日期{date}")
    if campus:
        parts.append(f"校区{campus}")

    return "，".join(parts) + "。"


def _build_structured_entry_for_group(rows: List[tuple], headers: HeaderMap) -> Dict[str, Any]:
    """将同一"序号"下的行转为结构化字典，供LLM生成自然语言。"""
    first = rows[0]
    
    # 根据业务大类确定结构
    business_type = _get(first, headers, "业务大类") or _get(first, headers, "业务类型") or "报销业务"
    
    entry: Dict[str, Any] = {
        "businessType": business_type,
        "login": {
            "username": _get(first, headers, "账号") or _get(first, headers, "用户名"),
            "password": _get(first, headers, "密码"),
        },
        "project": {
            "projectNumber": _get(first, headers, "报销项目号") or _get(first, headers, "项目号"),
            "attachmentCount": _normalize_number(_get(first, headers, "附件张数") or _get(first, headers, "附件")),
            "paymentMethod": _get(first, headers, "支付方式"),
            "remarks": _get(first, headers, "备注"),
            "special": _get(first, headers, "特殊事项说明"),
        },
        "expenses": [],
        "personnel": [],
        "travelPerson": [],
        "travelExpenses": [],
        "appointment": {
            "date": _get(first, headers, "日期"),
            "campus": (_get(first, headers, "校区") or "").replace("$$", ""),
        },
    }

    # 汇总科目（适用于报销业务）
    for r in rows:
        subject = _get(r, headers, "科目")
        amt = _normalize_number(_get(r, headers, "金额"))
        if subject or amt:
            entry["expenses"].append({"category": subject, "amount": amt})

    # 汇总转卡信息（适用于报销业务）
    for r in rows:
        emp_id = _get(r, headers, "转卡信息工号") or _get(r, headers, "工号")
        card_tail = _get(r, headers, "卡号尾号").lstrip("*")
        personal_amount = _normalize_number(_get(r, headers, "个人金额"))
        if emp_id or card_tail or personal_amount:
            entry["personnel"].append({
                "ID": emp_id,
                "bankCard": card_tail,
                "amount": personal_amount,
            })

    # 汇总差旅人员信息（适用于差旅业务）
    for r in rows:
        travel_id = _get(r, headers, "出差人")
        name = _get(r, headers, "姓名")
        person_type = _get(r, headers, "人员类型")
        if travel_id or name:
            entry["travelPerson"].append({
                "ID": travel_id,
                "name": name,
                "personType": person_type,
                "workUnit": "电子科技大学",  # 默认工作单位
                "title": "无",  # 默认职称
            })

    # 汇总差旅费用信息（适用于差旅业务）
    for r in rows:
        province = _get(r, headers, "省份")
        start_time = _get(r, headers, "起")
        end_time = _get(r, headers, "迄")
        airfare = _normalize_number(_get(r, headers, "飞机票"))
        trainfare = _normalize_number(_get(r, headers, "火车票"))
        other_transport = _normalize_number(_get(r, headers, "其他交通费"))
        accommodation = _normalize_number(_get(r, headers, "住宿费"))
        meal_arranged = _get(r, headers, "是否安排伙食")
        transport_arranged = _get(r, headers, "是否安排交通")
        transport_subsidy = _normalize_number(_get(r, headers, "交通补贴"))
        
        if province or start_time or end_time or airfare or trainfare or other_transport or accommodation:
            entry["travelExpenses"].append({
                "province": province,
                "startTime": start_time,
                "endTime": end_time,
                "airfare": airfare,
                "trainfare": trainfare,
                "otherTransport": other_transport,
                "accommodation": accommodation,
                "mealArranged": "true" if meal_arranged and "已安排" in meal_arranged else "false",
                "transportArranged": "true" if transport_arranged and "已安排" in transport_arranged else "false",
                "transportSubsidy": transport_subsidy,
            })

    return entry


def _build_llm_prompt(entry: Dict[str, Any]) -> str:
    """构造提示词，请Qwen用中文生成一段精炼的报销自然语言总结。"""
    business_type = entry.get("businessType", "")
    
    # 根据业务类型定制规则
    if "差旅" in business_type or "出差" in business_type:
        rules = (
            "你是财务自动化系统的助手。请根据给定的结构化数据，生成一段中文自然语言总结，要求：\n"
            "- 只输出一段自然语言，不要输出JSON或其它多余说明；\n"
            "- 若存在login.username或login.password，请在句首加入：登录账号{login.username}，密码{login.password}；\n"
            "- 按业务口语化顺序组织：业务大类、报销项目号、附件张数（加'张'）、备注、特殊事项、出差人员信息、差旅费用信息、转卡信息、预约日期与地点；\n"
            "- 将业务大类一项，单独形成一句话，如：业务大类为业务出差旅费。；\n"
            "- 对列表类信息进行顺序编号：\n"
            "  · 出差人员：按出现顺序标注为'出差人员1:姓名柳阳，工号2021090912003，人员类型其他人员；出差人员2:姓名刘宇航，工号202422090504，人员类型其他人员'（示例）；\n"
            "  · 差旅费用：按出现顺序标注为'差旅费用1:省份2，出差地点电子科技大学，起始时间2024-12-26，结束时间2024-12-27，其他交通费36，未安排伙食，未安排交通；差旅费用2:...'（示例）；\n"
            "  · 转卡信息：按出现顺序标注为'工号1:202512091010，卡号尾号5216；工号2:202422090504，卡号尾号无'（示例）；\n"
            "- 保持金额与附件张数为数字，自动补齐'张'等量词；\n"
            "- 不要遗漏任何非空字段；\n"
            "- 语句以句号结束。\n"
        )
    else:
        rules = (
            "你是财务自动化系统的助手。请根据给定的结构化数据，生成一段中文自然语言总结，要求：\n"
            "- 只输出一段自然语言，不要输出JSON或其它多余说明；\n"
            "- 若存在login.username或login.password，请在句首加入：登录账号{login.username}，密码{login.password}；\n"
            "- 按业务口语化顺序组织：业务大类、报销项目号、附件张数（加'张'）、支付方式、备注、特殊事项、科目与金额、转卡信息、预约日期与地点；\n"
            "- 将业务大类一项，单独形成一句话，如：业务大类为报销业务。；\n"
            "- 对列表类信息进行顺序编号：\n"
            "  · 科目与金额：按出现顺序标注为'科目1电费100、科目2专利费200'（示例）；\n"
            "  · 转卡信息：按出现顺序标注为'工号1:5070016，卡号尾号2818，个人金额50；工号2:202422090507，卡号尾号5054，个人金额50'（示例）；\n"
            "- 保持金额与附件张数为数字，自动补齐'张'等量词；\n"
            "- 不要遗漏任何非空字段；\n"
            "- 语句以句号结束。\n"
        )
    
    payload = {
        "businessType": entry.get("businessType"),
        "login": entry.get("login"),
        "project": entry.get("project"),
        "expenses": entry.get("expenses"),
        "personnel": entry.get("personnel"),
        "travelPerson": entry.get("travelPerson"),
        "travelExpenses": entry.get("travelExpenses"),
        "appointment": entry.get("appointment"),
    }
    return (
        rules
        + "\n结构化数据如下（JSON）：\n"
        + json.dumps(payload, ensure_ascii=False)
    )


def generate_nl_from_excel_via_llm(filepath: str, sheet_name: Optional[str] = None) -> List[str]:
    """通过本地Qwen（Ollama）对每个序号分组生成自然语言总结。"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = _read_sheet_headers(ws)
    data_rows: List[tuple] = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    groups = _group_rows_by_serial(data_rows, headers)

    ollama_base = os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434")
    model = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")

    results: List[str] = []
    for serial in sorted(groups.keys(), key=lambda x: str(x)):
        entry = _build_structured_entry_for_group(groups[serial], headers)
        prompt = _build_llm_prompt(entry)
        try:
            resp = requests.post(
                f"{ollama_base}/api/generate",
                json={"model": model, "prompt": prompt, "stream": False, "options": {"temperature": 0.2}},
                timeout=60,
            )
            if resp.status_code == 200:
                data = resp.json()
                text = (data.get("response") or "").strip()
                # 兜底：若空，则退回本地模板
                results.append(text or _build_summary_for_group(groups[serial], headers))
            else:
                results.append(_build_summary_for_group(groups[serial], headers))
        except Exception:
            # LLM不可用时，退回本地模板
            results.append(_build_summary_for_group(groups[serial], headers))
    return results


def generate_nl_from_excel(filepath: str, sheet_name: Optional[str] = None) -> List[str]:
    """从Excel文件读取数据，按“序号”分组，生成每个分组的自然语言。"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = _read_sheet_headers(ws)
    data_rows: List[tuple] = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    groups = _group_rows_by_serial(data_rows, headers)

    results: List[str] = []
    for serial in sorted(groups.keys(), key=lambda x: str(x)):
        summary = _build_summary_for_group(groups[serial], headers)
        results.append(summary)
    return results


def generate_single_nl_from_excel(
    filepath: str,
    sheet_name: Optional[str],
    serial: str | int,
    use_llm: bool = True,
) -> str:
    """按指定序号返回该组的自然语言总结。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名（可为 None，表示激活表）
        serial: 目标序号（字符串或整数，按表中“序号”列匹配）
        use_llm: 为 True 时调用本地 Qwen 生成；失败则回退到本地模板

    Returns:
        该序号的自然语言总结（未找到则返回空字符串）
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = _read_sheet_headers(ws)
    data_rows: List[tuple] = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    groups = _group_rows_by_serial(data_rows, headers)

    target_key = str(serial)
    if target_key not in groups:
        # 有些表头值可能为数字，我们再尝试宽松匹配（去掉前后空格）
        alt = {str(k).strip(): v for k, v in groups.items()}
        if target_key.strip() not in alt:
            return ""
        rows = alt[target_key.strip()]
    else:
        rows = groups[target_key]

    if use_llm:
        # 调用LLM生成；失败则回退
        entry = _build_structured_entry_for_group(rows, headers)
        prompt = _build_llm_prompt(entry)
        try:
            ollama_base = os.environ.get("OLLAMA_BASE_URL", "http://localhost:11434")
            model = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")
            resp = requests.post(
                f"{ollama_base}/api/generate",
                json={"model": model, "prompt": prompt, "stream": False, "options": {"temperature": 0.2}},
                timeout=60,
            )
            if resp.status_code == 200:
                data = resp.json()
                text = (data.get("response") or "").strip()
                return text or _build_summary_for_group(rows, headers)
        except Exception:
            pass
    # 回退到本地模板
    return _build_summary_for_group(rows, headers)


def main() -> int:
    # 直接指定文件路径和工作表名
    path = r"C:\Users\FH\source\repos\Auto Finan\Auto Finan\420财务050823.xlsx"
    sheet = "3-差旅"
    
    print(f"正在读取Excel文件: {path}")
    print(f"工作表名: {sheet}")
    print("=" * 50)
    
    try:
        # 优先通过本地Qwen生成；失败则退回本地模板
        texts = generate_nl_from_excel_via_llm(path, sheet)
        print("=== 由Excel生成的报销自然语言（LLM）===")
        for idx, t in enumerate(texts, 1):
            print(f"{idx}. {t}")
        return 0
    except FileNotFoundError:
        print(f"错误: 文件未找到 - {path}")
        return 1
    except Exception as e:
        print(f"处理Excel文件时发生错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())


