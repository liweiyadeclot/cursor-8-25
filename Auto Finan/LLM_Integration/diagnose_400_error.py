#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
400 错误诊断工具

帮助诊断为什么会出现 400 错误
"""

import os
import sys
from openpyxl import load_workbook

# 修复 Windows 控制台编码
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except:
        pass


def diagnose(excel_path, sheet_name, serial):
    """诊断问题"""
    print("=" * 80)
    print("400 错误诊断")
    print("=" * 80)
    print()
    
    # 1. 检查文件是否存在
    print("1. 检查 Excel 文件...")
    if not os.path.exists(excel_path):
        print(f"   ❌ 文件不存在: {excel_path}")
        print(f"   💡 请检查文件路径是否正确")
        return False
    else:
        file_size = os.path.getsize(excel_path)
        print(f"   ✅ 文件存在")
        print(f"   路径: {excel_path}")
        print(f"   大小: {file_size / 1024:.2f} KB")
    
    # 2. 检查工作表
    print(f"\n2. 检查工作表 '{sheet_name}'...")
    try:
        wb = load_workbook(excel_path, data_only=True)
        available_sheets = wb.sheetnames
        print(f"   ✅ 文件可读取")
        print(f"   所有工作表: {available_sheets}")
        
        if sheet_name not in available_sheets:
            print(f"   ❌ 工作表 '{sheet_name}' 不存在")
            print(f"   💡 可用工作表: {', '.join(available_sheets)}")
            return False
        else:
            print(f"   ✅ 工作表 '{sheet_name}' 存在")
            ws = wb[sheet_name]
    except Exception as e:
        print(f"   ❌ 无法读取文件: {e}")
        return False
    
    # 3. 检查表头
    print(f"\n3. 检查表头...")
    try:
        headers = {}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, val in enumerate(header_row):
            key = str(val or "").strip()
            if key:
                headers[key] = idx
        
        print(f"   ✅ 表头读取成功")
        print(f"   关键列: 序号={headers.get('序号', '未找到')}, "
              f"账号={headers.get('账号', headers.get('登录界面工号', '未找到'))}, "
              f"业务大类={headers.get('业务大类', headers.get('选择业务大类', '未找到'))}")
    except Exception as e:
        print(f"   ❌ 读取表头失败: {e}")
        return False
    
    # 4. 检查序号
    print(f"\n4. 检查序号 '{serial}'...")
    try:
        from collections import defaultdict
        groups = defaultdict(list)
        current_serial = None
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            serial_idx = headers.get("序号")
            if serial_idx is not None:
                serial_val = row[serial_idx]
                if serial_val is not None:
                    current_serial = str(serial_val).strip()
            if current_serial:
                groups[current_serial].append(row)
        
        all_serials = sorted(set(groups.keys()), key=lambda x: int(x) if x.isdigit() else 0)
        print(f"   ✅ 找到 {len(all_serials)} 个序号")
        print(f"   所有序号: {', '.join(all_serials[:10])}{'...' if len(all_serials) > 10 else ''}")
        
        target_serial = str(serial).strip()
        if target_serial not in groups:
            print(f"   ❌ 序号 '{serial}' 不存在")
            print(f"   💡 可用序号: {', '.join(all_serials[:10])}")
            return False
        else:
            rows = groups[target_serial]
            print(f"   ✅ 序号 '{serial}' 存在，包含 {len(rows)} 行数据")
            
            # 检查数据完整性
            if rows:
                first_row = rows[0]
                has_login = bool(headers.get("账号") or headers.get("登录界面工号"))
                has_business_type = bool(headers.get("业务大类") or headers.get("选择业务大类"))
                has_project = bool(headers.get("项目号") or headers.get("报销项目号"))
                
                print(f"   数据检查:")
                print(f"     登录信息: {'✅' if has_login else '❌'}")
                print(f"     业务大类: {'✅' if has_business_type else '❌'}")
                print(f"     项目信息: {'✅' if has_project else '❌'}")
    except Exception as e:
        print(f"   ❌ 检查序号失败: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    print(f"\n✅ 所有检查通过！应该可以正常生成提示词")
    return True


if __name__ == "__main__":
    # 默认测试数据
    excel_path = r"C:\Users\FH\source\repos\Auto Finan\Auto Finan\420财务050823.xlsx"
    sheet_name = "3-报销"
    serial = "1"
    
    # 可以从命令行参数获取
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    if len(sys.argv) > 2:
        sheet_name = sys.argv[2]
    if len(sys.argv) > 3:
        serial = sys.argv[3]
    
    success = diagnose(excel_path, sheet_name, serial)
    
    print("\n" + "=" * 80)
    if success:
        print("✅ 诊断完成，未发现问题")
        print("💡 如果仍然出现 400 错误，请检查服务日志")
    else:
        print("❌ 发现问题，请根据上述提示修复")
    print("=" * 80)
    
    sys.exit(0 if success else 1)

